# -*- coding:utf-8 -*-
import datetime
import openpyxl
import os
import settings

Seat = []  # リスト型で保存

No_Vacant_Seat = settings.No_Vacant_Seat


def time():
    """時間の取得関数"""
    # データタイムの現在の値
    dt_now = datetime.datetime.now()
    # 今の何年何月何日の値の取得
    day_today = dt_now.strftime('%Y%m%d')
    # 今の時間の取得
    time_now = dt_now.strftime('%H:%M:%S')
    return day_today, time_now


def month_day():
    """何年何月何日の取得関数"""
    # データタイムの現在の値
    dt_now = datetime.datetime.now()
    # 年と月の同時取得
    day_yearmonth = dt_now.strftime('%Y%m')
    # 月の取得
    day_month = dt_now.strftime('%m')
    # 日の取得
    day_day = dt_now.strftime('%d')
    return day_yearmonth, day_month, day_day


def New_day_book():
    day_today, time_now = time()
    day_yearmonth, day_month, day_day = month_day()
    if not os.path.isfile(f'../log/Seat_{day_yearmonth}.xlsx'):
        book = openpyxl.Workbook(f'../log/Seat_{day_yearmonth}.xlsx')
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()
        book = openpyxl.load_workbook(f'../log/Seat_{day_yearmonth}.xlsx')
        active_sheet = book.worksheets[0]
        active_sheet.title = f"{day_month}月{day_day}日"
        active_sheet.append(["名前", "席番号", "日付", "入室時間", "退室時間", "学年", "コース"])
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()


def Open_book():
    day_today, time_now = time()
    day_yearmonth, day_month, day_day = month_day()
    if os.path.isfile(f'../log/Seat_{day_yearmonth}.xlsx'):
        book = openpyxl.load_workbook(f'../log/Seat_{day_yearmonth}.xlsx')
        if not f"{day_month}月{day_day}日" in book.sheetnames:
            book.create_sheet(title=f"{day_month}月{day_day}日")
            active_sheet = book[f"{day_month}月{day_day}日"]
            active_sheet.append(
                ["名前", "席番号", "日付", "入室時間", "退室時間", "学年", "コース"])
            book.save(f'../log/Seat_{day_yearmonth}.xlsx')

        active_sheet = book[f"{day_month}月{day_day}日"]
        size = len(Seat)
        t = 0
        for row in active_sheet.rows:
            if row[0].row == 1:
                # １行目
                header_cells = row
            else:
                # ２行目以降
                row_dic = {}
                # セルの値を「key-value」で登録
                for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                    row_dic[k.value] = v.value
                if t < size:
                    Seat[t] = row_dic
                    t = t + 1
                else:
                    Seat.append(row_dic)
        
        for i in Seat:
            if (i["退室時間"] == None) and (not i["席番号"] in No_Vacant_Seat):
                No_Vacant_Seat.append(i["席番号"])
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()
    return No_Vacant_Seat


def append_time_in(name, number, year, course):  # 学年, コース追加
    day_today, time_now = time()
    day_yearmonth, day_month, day_day = month_day()
    if os.path.isfile(f'../log/Seat_{day_yearmonth}.xlsx'):
        book = openpyxl.load_workbook(
            filename=f'../log/Seat_{day_yearmonth}.xlsx')
        active_sheet = book[f"{day_month}月{day_day}日"]
        active_sheet.append(
            [name, number, day_today, time_now, "", year, course])  # 学年, コース追加
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()


def leave_seat_time(name, number):
    day_today, time_now = time()
    day_yearmonth, day_month, day_day = month_day()
    if os.path.isfile(f'../log/Seat_{day_yearmonth}.xlsx'):
        book = openpyxl.load_workbook(
            filename=f'../log/Seat_{day_yearmonth}.xlsx')
        active_sheet = book[f"{day_month}月{day_day}日"]
        size = len(Seat)
        t = 0
        for row in active_sheet.rows:
            if row[0].row == 1:
                # １行目
                header_cells = row
            else:
                # ２行目以降
                row_dic = {}
                # セルの値を「key-value」で登録
                for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                    row_dic[k.value] = v.value
                if t < size:
                    Seat[t] = row_dic
                    t = t + 1
                else:
                    Seat.append(row_dic)

        for i in Seat:
            if i["席番号"] == number:
                if i["名前"] == name:
                    if i["退室時間"] is None:
                        active_sheet.cell(
                            column=5, row=Seat.index(i) + 2, value=time_now)
                        tdatetime1 = datetime.datetime.strptime(
                            time_now, '%H:%M:%S')
                        tdatetime2 = datetime.datetime.strptime(
                            i["入室時間"], '%H:%M:%S')
                        String_time = ""
                        total = tdatetime1 - tdatetime2
                        total = total.seconds
                        if total / 3600 >= 1:
                            t = total // 3600
                            String_time = String_time + f"{t} 時間"
                            total = total - t * 3600
                        String_time = String_time + f"{total // 60} 分"
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()
        return String_time


def give_name(number):
    day_today, time_now = time()
    day_yearmonth, day_month, day_day = month_day()
    if os.path.isfile(f'../log/Seat_{day_yearmonth}.xlsx'):
        book = openpyxl.load_workbook(
            filename=f'../log/Seat_{day_yearmonth}.xlsx')
        active_sheet = book[f"{day_month}月{day_day}日"]
        size = len(Seat)
        t = 0
        for row in active_sheet.rows:
            if row[0].row == 1:
                # １行目
                header_cells = row
            else:
                # ２行目以降
                row_dic = {}
                # セルの値を「key-value」で登録
                for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                    row_dic[k.value] = v.value
                if t < size:
                    Seat[t] = row_dic
                    t = t + 1
                else:
                    Seat.append(row_dic)
        book.save(f'../log/Seat_{day_yearmonth}.xlsx')
        book.close()

        for i in Seat:
            if i["席番号"] == number:
                if i["退室時間"] is None:
                    Student = i["名前"]
                    return Student
