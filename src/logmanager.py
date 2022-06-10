import datetime
import openpyxl
import os


class LogManager():
    def __init__(self):
        self.dict_DAY = {}
        self.LOGFILE = ""

        # データを入れる
        self.SEAT = []
        self.Not_Available_Seat = []
        self.time()
        pass

    def time(self) -> None:
        """時間の取得関数"""
        # データタイムの現在の値
        dt_now = datetime.datetime.now()
        self.dict_DAY['Year'] = dt_now.strftime('%Y')
        self.dict_DAY['Month'] = dt_now.strftime('%m')
        self.dict_DAY['Day'] = dt_now.strftime('%d')
        self.dict_DAY['Hour'] = dt_now.strftime('%H')
        self.dict_DAY['Minute'] = dt_now.strftime('%M')
        self.dict_DAY['Second'] = dt_now.strftime('%S')
        self.LOGFILE = f"../log/Seat_{self.dict_DAY['Year']}{self.dict_DAY['Month']}.xlsx"
        pass

    def Create(self) -> None:
        self.time()
        if not os.path.isfile(self.LOGFILE):
            book = openpyxl.Workbook(self.LOGFILE)
            book.save(self.LOGFILE)
            book.close()
            book = openpyxl.load_workbook(self.LOGFILE)
            active_sheet = book.worksheets[0]
            active_sheet.title = f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"
            active_sheet.append(
                ["名前", "席番号", "日付", "入室時間", "退室時間", "学年", "コース"])
            book.save(self.LOGFILE)
            book.close()
        pass

    def Open(self) -> list:
        self.time()
        if os.path.isfile(self.LOGFILE):
            book = openpyxl.load_workbook(self.LOGFILE)
            if not (f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日" in book.sheetnames):
                book.create_sheet(
                    title=f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日")
                active_sheet = book[f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"]
                active_sheet.append(
                    ["名前", "席番号", "日付", "入室時間", "退室時間", "学年", "コース"])
                book.save(self.LOGFILE)

            active_sheet = book[f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"]

            size = len(self.SEAT)

            t = 0

            for row in active_sheet.rows:
                if row[0].row == 1:
                    # １行目
                    header_cells = row
                else:
                    # ２行目以降
                    row_dic = {}
                    # セルの値を「key-value」で登録
                    for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                        row_dic[k.value] = v.value
                    if t < size:  # 何の条件分岐？
                        self.SEAT[t] = row_dic
                        t = t + 1
                    else:
                        self.SEAT.append(row_dic)
            self.Not_Available_Seat = []

            for i in self.SEAT:
                if (i["退室時間"] is None) and not (i["席番号"] in self.Not_Available_Seat):
                    self.Not_Available_Seat.append(i["席番号"])
            book.save(self.LOGFILE)
            book.close()
        return self.Not_Available_Seat

    def LOG_Append(self, name, number, year, course) -> None:  # 学年, コース追加
        self.time()
        if os.path.isfile(self.LOGFILE):
            book = openpyxl.load_workbook(filename=self.LOGFILE)
            active_sheet = book[f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"]
            active_sheet.append(
                [
                    name, number,
                    f"{self.dict_DAY['Year']}{self.dict_DAY['Month']}{self.dict_DAY['Day']}",
                    f"{self.dict_DAY['Hour']}:{self.dict_DAY['Minute']}:{self.dict_DAY['Second']}",
                    "",
                    year,
                    course
                ])  # 学年, コース追加
            book.save(self.LOGFILE)
            book.close()
        pass

    def LOG_Leave(self, name, number):
        self.time()
        if os.path.isfile(self.LOGFILE):
            book = openpyxl.load_workbook(filename=self.LOGFILE)
            active_sheet = book[f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"]
            size = len(self.SEAT)
            t = 0
            for row in active_sheet.rows:
                if row[0].row == 1:
                    # １行目
                    header_cells = row
                else:
                    # ２行目以降
                    row_dic = {}
                    # セルの値を「key-value」で登録
                    for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                        row_dic[k.value] = v.value
                    if t < size:
                        self.SEAT[t] = row_dic
                        t = t + 1
                    else:
                        self.SEAT.append(row_dic)

            for i in self.SEAT:
                if i["席番号"] == number:
                    if i["名前"] == name:
                        if i["退室時間"] is None:
                            active_sheet.cell(
                                column=5, row=self.SEAT.index(i) + 2, value=f"{self.dict_DAY['Hour']}:{self.dict_DAY['Minute']}:{self.dict_DAY['Second']}")
                            tdatetime1 = datetime.datetime.strptime(
                                f"{self.dict_DAY['Hour']}:{self.dict_DAY['Minute']}:{self.dict_DAY['Second']}", '%H:%M:%S')
                            tdatetime2 = datetime.datetime.strptime(
                                i["入室時間"], '%H:%M:%S')
                            STUDY_TIME = ""
                            total = tdatetime1 - tdatetime2
                            total = total.seconds
                            if total / 3600 >= 1:
                                t = total // 3600
                                STUDY_TIME = STUDY_TIME + f"{t} 時間"
                                total = total - t * 3600
                            STUDY_TIME = STUDY_TIME + f"{total // 60} 分"
            book.save(self.LOGFILE)
            book.close()
            return STUDY_TIME

    def give_name(self, number):
        self.time()
        if os.path.isfile(self.LOGFILE):
            book = openpyxl.load_workbook(filename=self.LOGFILE)
            active_sheet = book[f"{self.dict_DAY['Month']}月{self.dict_DAY['Day']}日"]
            size = len(self.SEAT)
            t = 0
            for row in active_sheet.rows:
                if row[0].row == 1:
                    # １行目
                    header_cells = row
                else:
                    # ２行目以降
                    row_dic = {}
                    # セルの値を「key-value」で登録
                    for k, v in zip(header_cells, row):  # zip 関数(forループで複数のリストの要素を取得)
                        row_dic[k.value] = v.value
                    if t < size:
                        self.SEAT[t] = row_dic
                        t = t + 1
                    else:
                        self.SEAT.append(row_dic)
            book.save(self.LOGFILE)
            book.close()

            for i in self.SEAT:
                if i["席番号"] == number:
                    if not i["退室時間"]:
                        Student = i["名前"]
                        return Student
